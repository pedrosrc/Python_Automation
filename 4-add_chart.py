from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

# Acessando area de trabalho e planilha
workbook = load_workbook("data/pivot_table.xlsx")
sheet = workbook["Relatório"]

# Referências das linhas e colunas
min_column = workbook.active.min_column
max_column = workbook.active.max_column
min_row = workbook.active.min_row
max_row = workbook.active.max_row

# Adicinando Dados E Categorias no Gráfico

barchart = BarChart()

data = Reference(
    sheet,
    min_col= min_column + 1,
    max_col= max_column,
    min_row= min_row,
    max_row= max_row
)

categories = Reference(
    sheet,
    min_col= min_column,
    max_col= min_column,        
    min_row= min_row + 1,
    max_row= max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#Criando o Gráfico

sheet.add_chart(barchart, "B10")
barchart.title = "Vendas por Fabricantes"
barchart.style = 2

#Salvando o Workbook
workbook.save("data/barchart.xlsx")